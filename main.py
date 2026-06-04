from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import os
import tkinter as tk
from tkinter import filedialog, messagebox

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def merge_tables(file1, file2, output):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    ws1 = wb1.active
    ws2 = wb2.active

    # 读表1列宽
    col_widths = {}
    for col_idx in range(1, ws1.max_column + 1):
        letter = get_column_letter(col_idx)
        col_widths[letter] = ws1.column_dimensions[letter].width

    # 表头样式
    header_styles = {}
    for col_idx in range(1, ws1.max_column + 1):
        cell = ws1.cell(1, col_idx)
        header_styles[col_idx] = {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment),
            'border': copy(cell.border),
            'number_format': cell.number_format,
        }

    # 表2索引
    ws2_index = {}
    for row in range(2, ws2.max_row + 1):
        code = ws2.cell(row, 1).value
        if code is not None:
            ws2_index[code] = row

    merge_cols = [3, 4, 5]  # 单子金额、上货金额、原因

    # 收集所有行数据
    all_rows = []
    ws1_keys = set()

    for row in range(2, ws1.max_row + 1):
        code = ws1.cell(row, 1).value
        if code is None:
            continue
        ws1_keys.add(code)

        row_data = {'code': code}
        for col_idx in range(1, ws1.max_column + 1):
            row_data[col_idx] = {
                'value': ws1.cell(row, col_idx).value,
                'font': copy(ws1.cell(row, col_idx).font),
                'alignment': copy(ws1.cell(row, col_idx).alignment),
                'number_format': ws1.cell(row, col_idx).number_format,
            }

        # 合并表2
        if code in ws2_index:
            r2 = ws2_index[code]
            for col_idx in merge_cols:
                v1 = ws1.cell(row, col_idx).value
                v2 = ws2.cell(r2, col_idx).value

                if v1 is None and v2 is None:
                    continue
                if v1 is None:
                    row_data[col_idx]['value'] = v2
                    continue
                if v2 is None:
                    row_data[col_idx]['value'] = v1
                    continue
                row_data[col_idx]['value'] = f'{v1}+{v2}'

        all_rows.append(row_data)

    # 表2独有的
    for row in range(2, ws2.max_row + 1):
        code = ws2.cell(row, 1).value
        if code is None or code in ws1_keys:
            continue
        row_data = {'code': code}
        for col_idx in range(1, ws2.max_column + 1):
            row_data[col_idx] = {
                'value': ws2.cell(row, col_idx).value,
                'font': copy(ws2.cell(row, col_idx).font),
                'alignment': copy(ws2.cell(row, col_idx).alignment),
                'number_format': ws2.cell(row, col_idx).number_format,
            }
        all_rows.append(row_data)

    # 按供应商编码排序
    all_rows.sort(key=lambda x: str(x['code']))

    # 写入输出
    wb_out = Workbook()
    ws_out = wb_out.active

    # 写表头
    for col_idx in range(1, ws1.max_column + 1):
        src = ws1.cell(1, col_idx)
        dst = ws_out.cell(1, col_idx, value=src.value)
        if col_idx in header_styles:
            s = header_styles[col_idx]
            dst.font = copy(s['font'])
            dst.fill = copy(s['fill'])
            dst.alignment = copy(s['alignment'])
            dst.number_format = s['number_format']
        dst.border = thin_border

    # 写数据行
    for i, row_data in enumerate(all_rows):
        out_row = i + 2
        for col_idx in range(1, ws1.max_column + 1):
            cell_data = row_data.get(col_idx, {})
            dst = ws_out.cell(out_row, col_idx)
            dst.value = cell_data.get('value')
            if 'font' in cell_data:
                dst.font = cell_data['font']
            if 'alignment' in cell_data:
                dst.alignment = cell_data['alignment']
            if 'number_format' in cell_data:
                dst.number_format = cell_data['number_format']
            dst.border = thin_border

    # 设置列宽
    for letter, width in col_widths.items():
        if width:
            ws_out.column_dimensions[letter].width = width

    wb_out.save(output)
    return len(all_rows)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title('供应商表格整合工具')
        self.root.geometry('500x280')
        self.root.resizable(False, False)

        self.file1 = tk.StringVar()
        self.file2 = tk.StringVar()
        self.output = tk.StringVar(value=os.path.join(os.path.expanduser('~'), 'Desktop', '整合结果.xlsx'))

        tk.Label(root, text='供应商表格整合工具', font=('微软雅黑', 14, 'bold')).pack(pady=10)

        frame = tk.Frame(root)
        frame.pack(pady=5, padx=20, fill='x')

        tk.Label(frame, text='表1:').grid(row=0, column=0, sticky='w')
        tk.Entry(frame, textvariable=self.file1, width=40).grid(row=0, column=1, padx=5)
        tk.Button(frame, text='选择', command=lambda: self.browse(self.file1)).grid(row=0, column=2)

        tk.Label(frame, text='表2:').grid(row=1, column=0, sticky='w', pady=5)
        tk.Entry(frame, textvariable=self.file2, width=40).grid(row=1, column=1, padx=5)
        tk.Button(frame, text='选择', command=lambda: self.browse(self.file2)).grid(row=1, column=2)

        tk.Label(frame, text='输出:').grid(row=2, column=0, sticky='w')
        tk.Entry(frame, textvariable=self.output, width=40).grid(row=2, column=1, padx=5)
        tk.Button(frame, text='选择', command=lambda: self.save_as()).grid(row=2, column=2)

        tk.Button(root, text='开始整合', command=self.run, font=('微软雅黑', 11), bg='#4CAF50', fg='white', width=15).pack(pady=20)

    def browse(self, var):
        path = filedialog.askopenfilename(filetypes=[('Excel文件', '*.xlsx *.xls')])
        if path:
            var.set(path)

    def save_as(self):
        path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel文件', '*.xlsx')])
        if path:
            self.output.set(path)

    def run(self):
        f1, f2, out = self.file1.get(), self.file2.get(), self.output.get()
        if not f1 or not f2:
            messagebox.showerror('错误', '请选择两张表')
            return
        try:
            count = merge_tables(f1, f2, out)
            messagebox.showinfo('完成', f'整合完成！共 {count} 行\n已保存到: {out}')
        except Exception as e:
            messagebox.showerror('错误', str(e))


if __name__ == '__main__':
    root = tk.Tk()
    App(root)
    root.mainloop()
